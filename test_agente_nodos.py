"""
test_agente_nodos.py — evalúa el loop multi-ronda de decidir_y_actuar.

Modo completamente simulado: ninguna acción real se ejecuta en el sistema.
- _ejecutar_con_verificacion: mockeado (no corre código, devuelve True)
- enviar_whatsapp / enviar_archivo_whatsapp: stubs que solo registran la llamada
- Sin escritura de archivos ni interacción con el sistema de archivos
- Contactos ficticios (no depende de contactos.json)
- 50 pruebas, 1–4 nodos, tipos de archivo variados (.txt .csv .json .md .log)
"""
import os
import sys
import queue
import time
import json as _json
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
os.environ["STEM_DEBUG_TEXTO"] = "1"

import ia       # noqa: E402
import whatsapp # noqa: E402

# ── Compat shim ──────────────────────────────────────────────────────────────
# decidir_y_actuar fue reemplazado por _ejecutar_turno cuando se introdujo el
# Orchestrator (ia.py ya no lo expone). Arma un messages[] de un solo turno y
# llama al loop actual para no tocar el resto del test.
def _decidir_y_actuar_compat(texto, audio_q, rec):
    messages = [
        {"role": "system", "content": f"{ia._TOOLS_SYSTEM}\n\n{ia._get_rutas_contexto()}"},
        {"role": "user", "content": texto},
    ]
    return ia._ejecutar_turno(messages, texto, audio_q, rec)

ia.decidir_y_actuar = _decidir_y_actuar_compat

# Orchestrator.confirmar_con_usuario() usa input() en modo debug (no existía
# cuando se escribió este test) — auto-confirmar como el resto de las funciones
# de confirmación mockeadas abajo.
import builtins  # noqa: E402
builtins.input = lambda *a, **kw: "si"

# ── Patches ────────────────────────────────────────────────────────────────────
ia._escuchar_confirmacion_debug = lambda: "si"
ia.hablar_edge         = lambda texto, *a, **kw: print(f"   [TTS] {texto}")
ia._reproducir_oracion = lambda texto: print(f"   [TTS-oracion] {texto}")

# orchestrator.confirmar_con_usuario() llama _hablar_stem() (Cartesia real) para
# el "¿procedo?" — no existía cuando se escribió este test. Mockear para no
# reproducir audio real ~7-10s por cada acción.
import orchestrator as _orch  # noqa: E402
_orch._hablar_stem = lambda texto, *a, **kw: print(f"   [TTS-confirm] {texto}")

# Stub WA — firma nueva: enviar_whatsapp(envios: list[dict])
_wa_calls_this_test: list[dict] = []

def _stub_enviar_whatsapp(envios: list[dict]) -> bool:
    for e in envios:
        _wa_calls_this_test.append({"type": "msg",
                                     "contacto": e.get("contacto", ""),
                                     "mensaje":  e.get("mensaje",  "")})
        print(f"   [WA-MSG] {e.get('contacto')}: {str(e.get('mensaje',''))[:60]}")
    return True

def _stub_enviar_archivo_whatsapp(envios: list[dict]) -> bool:
    for e in envios:
        _wa_calls_this_test.append({"type": "file",
                                     "contacto": e.get("contacto", ""),
                                     "ruta":     e.get("archivo", "")})
        print(f"   [WA-FILE] {e.get('contacto')}: {e.get('archivo', '')}")
    return True

whatsapp.enviar_whatsapp         = _stub_enviar_whatsapp
whatsapp.enviar_archivo_whatsapp = _stub_enviar_archivo_whatsapp

# Gatekeeper tracker: cuenta cuántas veces bloqueó el cierre
_gk_blocks_this_test: int = 0
_gk_total_blocks:     int = 0

_orig_quedan_pendientes = ia._quedan_pendientes

def _mock_quedan_pendientes(peticion_original: str, messages: list) -> bool:
    global _gk_blocks_this_test, _gk_total_blocks
    result = _orig_quedan_pendientes(peticion_original, messages)
    if result:
        _gk_blocks_this_test += 1
        _gk_total_blocks     += 1
    return result

ia._quedan_pendientes = _mock_quedan_pendientes

# Mock exec: NO ejecuta código real, solo contabiliza
_tools_this_test:  list[str] = []
_rondas_this_test: int = 0
_exec_nodos_this_test: int = 0
_total_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

_orig_create = ia._client.chat.completions.create

def _mock_create(*args, **kwargs):
    global _rondas_this_test
    resp = _orig_create(*args, **kwargs)
    _rondas_this_test += 1
    if resp.choices and resp.choices[0].message.tool_calls:
        _tools_this_test.append(resp.choices[0].message.tool_calls[0].function.name)
    if resp.usage:
        _total_usage["prompt_tokens"]     += resp.usage.prompt_tokens
        _total_usage["completion_tokens"] += resp.usage.completion_tokens
        _total_usage["total_tokens"]      += resp.usage.total_tokens
    return resp

ia._client.chat.completions.create = _mock_create

def _mock_ejecutar(descripcion, codigo, audio_q, rec, youtube_query=None):
    """Simulación completa: nunca corre código real. Solo registra el nodo."""
    global _exec_nodos_this_test
    if youtube_query:
        print(f"   [YT-MOCK] búsqueda: {youtube_query}")
    else:
        print(f"   [EXEC-MOCK] {descripcion}")
    _exec_nodos_this_test += 1
    return True

ia._ejecutar_con_verificacion = _mock_ejecutar

# El camino real de "ejecutar_accion"/"buscar_y_abrir_youtube" ya no pasa por
# ia._ejecutar_con_verificacion (eso es del modo agente legacy) — pasa por
# tools.filesystem._ejecutar_silencioso y tools.youtube._abrir_en_brave desde
# el refactor a tools/. Sin mockear esto, exec() correría código real en el
# filesystem y se abriría Brave de verdad.
import tools.filesystem as _tf  # noqa: E402
import tools.youtube as _ty     # noqa: E402

def _mock_ejecutar_silencioso(descripcion, codigo, youtube_query=None):
    global _exec_nodos_this_test
    if youtube_query:
        print(f"   [YT-MOCK] búsqueda: {youtube_query}")
    else:
        print(f"   [EXEC-MOCK] {descripcion}")
    _exec_nodos_this_test += 1
    return True

_tf._ejecutar_silencioso = _mock_ejecutar_silencioso

def _mock_abrir_en_brave(youtube_query):
    global _exec_nodos_this_test
    print(f"   [YT-MOCK] búsqueda: {youtube_query}")
    _exec_nodos_this_test += 1
    return True

_ty._abrir_en_brave = _mock_abrir_en_brave

# ── Vocabularios ───────────────────────────────────────────────────────────────
CONTACTOS = [
    "Carlos", "Ana", "Luis", "Sofia", "Miguel",
    "Elena", "Pedro", "Laura", "Diego", "Valeria",
    "Rodrigo", "Camila", "Javier", "Natalia", "Andres",
]

EXTENSIONES = [".txt", ".csv", ".json", ".md", ".log", ".xml", ".cfg", ".ini"]

NOMBRES = [
    "reporte", "resumen", "agenda", "datos", "config",
    "notas", "plan", "inventario", "estadisticas", "log_dia",
    "presupuesto", "contrato", "acuerdos", "pendientes", "borrador",
    "ideas", "tareas", "cronograma", "objetivo", "esquema",
    "informe", "memo", "prioridades", "guia", "listado",
    "registro", "apuntes", "plantilla", "propuesta", "analisis",
    "minutas", "checklist", "seguimiento", "resultados", "reporte_mes",
]

CONTENIDOS = [
    "reunión programada para las 3pm",
    "lista de compras: leche, pan, huevos",
    "tareas pendientes del día de hoy",
    "sin novedades por el momento",
    "ventas del mes: 100 unidades",
    "objetivos de la semana en curso",
    "resumen de la llamada con el cliente",
    "presupuesto estimado para el proyecto",
    "pendientes urgentes antes del viernes",
    "cinco consejos de productividad",
    "notas de la reunión de equipo",
    "cronograma de actividades del mes",
    "errores detectados en el sistema",
    "acciones correctivas identificadas",
    "balance financiero del trimestre",
    "indicadores clave de desempeño",
    "agenda del próximo sprint",
    "retrospectiva del proyecto anterior",
    "decisiones tomadas en la sesión",
    "riesgos identificados y mitigaciones",
]

MENSAJES_WA = [
    "Recuerda la reunión de mañana",
    "¿Puedes revisar el documento que te envié?",
    "Todo listo para el viernes",
    "Confirma si puedes asistir",
    "Llego en veinte minutos",
    "El proyecto fue aprobado",
    "Necesito tu firma antes del lunes",
    "Te paso el resumen por aquí",
    "Avanza con el informe, ya tengo los datos",
    "Confirmado para las 4pm",
]

QUERIES_YT = [
    "música para concentrarse",
    "meditación guiada 10 minutos",
    "lo mejor del jazz 2024",
    "tutoriales de Python avanzado",
    "noticias tecnología hoy",
    "lofi hip hop estudio",
    "ejercicios de respiración",
    "cocina italiana recetas",
]

CARPETAS = ["el escritorio", "descargas", "documentos"]

_DUMMY_REC = type("DummyRec", (), {"Reset": lambda self: None})()
_DUMMY_Q   = queue.Queue()
# Threshold para detectar rondas agotadas: ia.py usa MAX_RONDAS=10 internamente,
# pero el gatekeeper agrega llamadas extra por bloqueo — usamos 14 para no falsos positivos.
MAX_RONDAS = 14


def _pick(pool: list, i: int, offset: int = 0):
    return pool[(i * 7 + offset) % len(pool)]

def _pick_n(pool: list, i: int, n: int, offset: int = 0) -> list:
    seen, result = set(), []
    for step in range(len(pool)):
        c = pool[(i * 7 + offset + step) % len(pool)]
        if c not in seen:
            seen.add(c); result.append(c)
        if len(result) == n:
            break
    return result


def _reset():
    global _tools_this_test, _rondas_this_test, _exec_nodos_this_test, _gk_blocks_this_test
    _tools_this_test      = []
    _rondas_this_test     = 0
    _exec_nodos_this_test = 0
    _gk_blocks_this_test  = 0
    _wa_calls_this_test.clear()


# ── Generador de peticiones ────────────────────────────────────────────────────
def generar_peticion(i: int) -> tuple[str, int]:
    """
    Genera una petición de 5 o 6 nodos para evaluar el batch de enviar_archivo_whatsapp
    y la regla de tareas múltiples. Solo 5 y 6 nodos — 1–4 ya están validados.

    Distribución (30 tests):
      i  0-14 → 5 nodos (15 tests, 4 categorías rotando)
      i 15-29 → 6 nodos (15 tests, 4 categorías rotando)
    """
    nodos = 5 if i < 15 else 6

    cat = i % 4

    noms  = _pick_n(NOMBRES,     i, 5, offset=0)
    exts  = _pick_n(EXTENSIONES, i, 5, offset=1)
    dirs  = _pick_n(CARPETAS,    i, 3, offset=2)
    conts = _pick_n(CONTENIDOS,  i, 5, offset=3)
    cons  = _pick_n(CONTACTOS,   i, 5, offset=4)
    msgs  = _pick_n(MENSAJES_WA, i, 4, offset=5)
    yt    = _pick(QUERIES_YT, i, offset=6)

    d0, d1, d2 = dirs[0], dirs[1 % 3], dirs[2 % 3]

    # ── 5 nodos ─────────────────────────────────────────────────────────────
    # Nodos esperados por categoría:
    # cat 0: exec×3 + WA_file×1 + WA_msg×1              (envío al final)
    # cat 1: exec×2 + WA_file×2(batch) + WA_msg×1       (batch file send)
    # cat 2: exec×1 + WA_file×1 + exec×1 + WA_msg×1 + exec×1  (interleaved)
    # cat 3: exec×2 + WA_file×2(batch) + YT×1           (batch + YouTube)
    if nodos == 5:
        if cat == 0:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"crea {noms[2]}{exts[2]} en {d2} con '{conts[2]}', "
                   f"envíaselo {noms[0]}{exts[0]} a {cons[0]} como archivo por WhatsApp, "
                   f"y envíale a {cons[1]} el mensaje '{msgs[0]}'")
        elif cat == 1:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"envíaselo {noms[0]}{exts[0]} a {cons[0]} y también "
                   f"{noms[1]}{exts[1]} a {cons[1]} como archivos por WhatsApp, "
                   f"y envíale a {cons[2]} el mensaje '{msgs[0]}'")
        elif cat == 2:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"envíaselo a {cons[0]} como archivo por WhatsApp, "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"envíale a {cons[1]} el mensaje '{msgs[0]}', "
                   f"y crea {noms[2]}{exts[2]} en {d2} con '{conts[2]}'")
        else:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"envíaselo {noms[0]}{exts[0]} a {cons[0]} y "
                   f"{noms[1]}{exts[1]} a {cons[1]} también como archivos por WhatsApp, "
                   f"y busca en YouTube '{yt}'")

    # ── 6 nodos ─────────────────────────────────────────────────────────────
    # cat 0: exec×3 + WA_file×2(batch) + WA_msg×1
    # cat 1: exec×2 + WA_file×2(batch) + exec×1 + WA_msg×1  (crea después del batch)
    # cat 2: exec×3 + WA_msg×3(batch en una llamada)
    # cat 3: exec×2 + WA_file×2(batch) + WA_msg×2(batch)
    else:
        if cat == 0:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"crea {noms[2]}{exts[2]} en {d2} con '{conts[2]}', "
                   f"envíaselo {noms[0]}{exts[0]} a {cons[0]} y "
                   f"{noms[1]}{exts[1]} a {cons[1]} como archivos por WhatsApp, "
                   f"y envíale a {cons[2]} el mensaje '{msgs[0]}'")
        elif cat == 1:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"envíaselos a {cons[0]} y {cons[1]} como archivos por WhatsApp, "
                   f"crea {noms[2]}{exts[2]} en {d2} con '{conts[2]}', "
                   f"y envíale a {cons[2]} el mensaje '{msgs[0]}'")
        elif cat == 2:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"crea {noms[2]}{exts[2]} en {d2} con '{conts[2]}', "
                   f"y envíales mensajes por WhatsApp: a {cons[0]}: '{msgs[0]}', "
                   f"a {cons[1]}: '{msgs[1]}' y a {cons[2]}: '{msgs[2]}'")
        else:
            txt = (f"Crea {noms[0]}{exts[0]} en {d0} con '{conts[0]}', "
                   f"crea {noms[1]}{exts[1]} en {d1} con '{conts[1]}', "
                   f"envíale {noms[0]}{exts[0]} a {cons[0]} y {noms[1]}{exts[1]} a {cons[1]} "
                   f"como archivos por WhatsApp, "
                   f"y envíales también mensajes: a {cons[2]}: '{msgs[0]}' "
                   f"y a {cons[3]}: '{msgs[1]}'")

    return txt, nodos


# ── Ejecución ──────────────────────────────────────────────────────────────────
results: list[dict] = []

TOTAL = 30
print(f"\n{'='*65}")
print(f"test_agente_nodos.py — {TOTAL} pruebas × MAX_RONDAS={MAX_RONDAS}")
print("Modo: simulado (sin acciones reales). Solo nodos 5 y 6.")
print(f"{'='*65}\n")

for i in range(TOTAL):
    texto, nodos_esp = generar_peticion(i)

    print(f"[{i+1:02d}/{TOTAL}] ({nodos_esp} nodo{'s' if nodos_esp>1 else ''}) {texto[:78]}...")
    _reset()
    error_str = ""

    try:
        ia.decidir_y_actuar(texto, _DUMMY_Q, _DUMMY_REC)
    except Exception as exc:
        error_str = str(exc)
        print(f"   ERROR: {exc}")

    wa_files = [c for c in _wa_calls_this_test if c["type"] == "file"]
    wa_msgs  = [c for c in _wa_calls_this_test if c["type"] == "msg"]
    # Mensajes WA en un solo enviar_whatsapp con varios ítems cuentan por separado
    nodos_completados = _exec_nodos_this_test + len(wa_files) + len(wa_msgs)

    agoto_rondas = _rondas_this_test >= MAX_RONDAS
    precision    = min(nodos_completados / nodos_esp, 1.0) if nodos_esp > 0 else 0.0

    gk_str = f" | gk={_gk_blocks_this_test}" if _gk_blocks_this_test else ""
    print(
        f"   tools={_tools_this_test} | "
        f"nodos={nodos_completados}/{nodos_esp} | "
        f"rondas={_rondas_this_test}{gk_str} | "
        f"prec={precision:.2f}"
    )

    results.append({
        "test_n":             i + 1,
        "texto_instruccion":  texto,
        "nodos_esperados":    nodos_esp,
        "nodos_completados":  nodos_completados,
        "tools_llamadas":     " → ".join(_tools_this_test),
        "rondas_usadas":      _rondas_this_test,
        "gk_bloqueos":        _gk_blocks_this_test,
        "agoto_rondas":       agoto_rondas,
        "error":              error_str,
        "precision_estimada": round(precision, 4),
    })

    time.sleep(0.3)

# ── Excel ──────────────────────────────────────────────────────────────────────
out_path = Path(__file__).parent / "test_resultados_nodos.xlsx"

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"

    headers = list(results[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True)
        cell.fill      = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for row in results:
        ws.append([row[h] for h in headers])

    for col_idx, header in enumerate(headers, 1):
        max_len = max(len(str(header)), max(len(str(r[header])) for r in results))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 80)

    wb.save(out_path)
    print(f"\n[OK] Excel guardado: {out_path}")

except Exception as exc:
    print(f"\n[WARN] No se guardó Excel: {exc}")
    for r in results:
        print(_json.dumps(r, ensure_ascii=True))

# ── Resumen consola ────────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print("RESUMEN FINAL")
print(f"{'='*65}")

by_nodos: dict[int, list[float]] = {}
for r in results:
    by_nodos.setdefault(r["nodos_esperados"], []).append(r["precision_estimada"])

prec_total = sum(r["precision_estimada"] for r in results) / len(results)
agotaron   = sum(1 for r in results if r["agoto_rondas"])
con_error  = sum(1 for r in results if r["error"])

print(f"Precisión promedio total:     {prec_total:.2%}")
for n in sorted(by_nodos):
    vals = by_nodos[n]
    label = f"{n} nodo{'s' if n>1 else ''}"
    print(f"  {label} ({len(vals):2d} prueba{'s' if len(vals)>1 else ''}):  {sum(vals)/len(vals):.2%}")
print(f"Pruebas que agotaron rondas:  {agotaron}/{len(results)}")
print(f"Pruebas con error:            {con_error}/{len(results)}")
gk_casos = sum(1 for r in results if r["gk_bloqueos"] > 0)
print(f"Activaciones del gatekeeper:  {_gk_total_blocks} bloqueos en {gk_casos} pruebas")

print(f"\n{'='*65}")
print("USO DE TOKENS")
print(f"{'='*65}")
print(f"  Entrada   (prompt_tokens):     {_total_usage['prompt_tokens']:>10,}")
print(f"  Salida    (completion_tokens): {_total_usage['completion_tokens']:>10,}")
print(f"  Total     (total_tokens):      {_total_usage['total_tokens']:>10,}")
print(f"{'='*65}\n")
